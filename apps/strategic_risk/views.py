from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from functools import wraps
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
import json
from .models import StrategicPlan, ExternalEnvironment, FinancialEnvironment, InternalDiagnosis, Perspectiva, ObjetivoEstrategico, Indicador, MetaPeriodo, StrategicMatrix, BusinessModelCanvas, CorporatePhilosophy, TipoObjetivo, AreaResponsable, ResponsablePlan
from .forms import StrategicPlanForm, ExternalEnvironmentForm, FinancialEnvironmentForm, InternalDiagnosisForm

def check_module_access(module_name):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_superuser:
                if module_name in getattr(request.user, 'hidden_modules', []):
                    return render(request, 'access_denied.html', {'module_name': module_name})
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

@login_required
@check_module_access('Estrategia y Objetivos')
def dashboard(request):
    planes = StrategicPlan.objects.all().order_by('-start_year')
    context = {
        'page_title': 'Planeamiento Estratégico - Dashboard',
        'planes': planes
    }
    return render(request, 'strategic_risk/dashboard.html', context)

@login_required
@check_module_access('Estrategia y Objetivos')
def plan_create(request):
    if request.method == 'POST':
        form = StrategicPlanForm(request.POST)
        if form.is_valid():
            plan = form.save(commit=False)
            plan.created_by = request.user
            plan.save()
            # Crear registros vacíos por defecto para entornos y diagnóstico
            ExternalEnvironment.objects.create(plan=plan)
            FinancialEnvironment.objects.create(plan=plan)
            InternalDiagnosis.objects.create(plan=plan)
            # Crear matrices vacías por defecto
            StrategicMatrix.objects.create(plan=plan, matrix_type='FODA')
            StrategicMatrix.objects.create(plan=plan, matrix_type='EFI')
            StrategicMatrix.objects.create(plan=plan, matrix_type='EFE')
            StrategicMatrix.objects.create(plan=plan, matrix_type='MPC')
            BusinessModelCanvas.objects.create(plan=plan)
            CorporatePhilosophy.objects.create(plan=plan)
            
            # Guardar Perspectivas
            nombres = request.POST.getlist('perspectivas[]')
            # Intentar obtener la organización del usuario para TenantAwareModel
            from users.models import Organization
            org = getattr(request.user, 'organization', None)
            if not org:
                org = Organization.objects.first()
                if not org:
                    org = Organization.objects.create(name="Organización Principal")
                
            for nombre in nombres:
                nombre = nombre.strip()
                if nombre:
                    Perspectiva.objects.create(plan=plan, nombre=nombre, organization=org)
            
            # Guardar Tipos de Objetivo
            tipos = request.POST.getlist('tipos_objetivo[]')
            for t in tipos:
                t = t.strip()
                if t: TipoObjetivo.objects.create(plan=plan, nombre=t, organization=org)

            # Guardar Áreas Responsables
            areas = request.POST.getlist('areas_responsables[]')
            for a in areas:
                a = a.strip()
                if a: AreaResponsable.objects.create(plan=plan, nombre=a, organization=org)

            # Guardar Responsables
            responsables = request.POST.getlist('responsables[]')
            for r in responsables:
                r = r.strip()
                if r: ResponsablePlan.objects.create(plan=plan, nombre=r, organization=org)
                    
            request.session['active_strategic_plan_id'] = plan.id
            messages.success(request, 'Plan Estratégico creado exitosamente.')
            return redirect('strategic_risk:dashboard')
    else:
        form = StrategicPlanForm()
    
    context = {
        'page_title': 'Crear Plan Estratégico',
        'form': form
    }
    return render(request, 'strategic_risk/plan_form.html', context)

@login_required
@check_module_access('Estrategia y Objetivos')
def plan_update(request, pk):
    plan = get_object_or_404(StrategicPlan, pk=pk)
    if request.method == 'POST':
        form = StrategicPlanForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            
            # Guardar Perspectivas
            nombres = request.POST.getlist('perspectivas[]')
            from users.models import Organization
            org = getattr(request.user, 'organization', None)
            if not org:
                org = Organization.objects.first()
                if not org:
                    org = Organization.objects.create(name="Organización Principal")
            
            existing_perspectivas = {p.nombre: p for p in plan.perspectivas.all()}
            keep_ids = []
            
            for nombre in nombres:
                nombre = nombre.strip()
                if not nombre: continue
                if nombre in existing_perspectivas:
                    keep_ids.append(existing_perspectivas[nombre].id)
                else:
                    p = Perspectiva.objects.create(plan=plan, nombre=nombre, organization=org)
                    keep_ids.append(p.id)
            
            # Eliminar las que fueron quitadas (esto borrará también Objetivos en cascada, si los hubiera)
            plan.perspectivas.exclude(id__in=keep_ids).delete()
            
            # Guardar Tipos
            tipos = request.POST.getlist('tipos_objetivo[]')
            plan.tipos_objetivo.all().delete()
            for t in tipos:
                t = t.strip()
                if t: TipoObjetivo.objects.create(plan=plan, nombre=t, organization=org)

            # Guardar Areas
            areas = request.POST.getlist('areas_responsables[]')
            plan.areas_responsables.all().delete()
            for a in areas:
                a = a.strip()
                if a: AreaResponsable.objects.create(plan=plan, nombre=a, organization=org)

            # Guardar Responsables
            responsables = request.POST.getlist('responsables[]')
            plan.responsables.all().delete()
            for r in responsables:
                r = r.strip()
                if r: ResponsablePlan.objects.create(plan=plan, nombre=r, organization=org)
            
            messages.success(request, 'Plan Estratégico actualizado exitosamente.')
            return redirect('strategic_risk:dashboard')
    else:
        form = StrategicPlanForm(instance=plan)
    
    perspectivas = plan.perspectivas.all()
    tipos_objetivo = plan.tipos_objetivo.all()
    areas_responsables = plan.areas_responsables.all()
    responsables = plan.responsables.all()

    context = {
        'page_title': 'Editar Plan Estratégico', 
        'form': form, 
        'plan': plan, 
        'perspectivas': perspectivas,
        'tipos_objetivo': tipos_objetivo,
        'areas_responsables': areas_responsables,
        'responsables': responsables
    }
    return render(request, 'strategic_risk/plan_form.html', context)

@login_required
@check_module_access('Estrategia y Objetivos')
def plan_copy(request, pk):
    original_plan = get_object_or_404(StrategicPlan, pk=pk)
    
    # 1. Copiar Plan Estratégico
    new_plan = StrategicPlan.objects.create(
        name=f"{original_plan.name} (Copia)",
        institution=original_plan.institution,
        start_year=original_plan.start_year,
        horizon_years=original_plan.horizon_years,
        status='DRAFT',
        version=original_plan.version,
        created_by=request.user
    )
    
    # 2. Copiar Entornos
    ext_env = getattr(original_plan, 'external_environment', None)
    if ext_env:
        ext_env.pk = None
        ext_env.plan = new_plan
        ext_env.save()
    else:
        ExternalEnvironment.objects.create(plan=new_plan)
        
    fin_env = getattr(original_plan, 'financial_environment', None)
    if fin_env:
        fin_env.pk = None
        fin_env.plan = new_plan
        fin_env.save()
    else:
        FinancialEnvironment.objects.create(plan=new_plan)
        
    int_diag = getattr(original_plan, 'internal_diagnosis', None)
    if int_diag:
        int_diag.pk = None
        int_diag.plan = new_plan
        int_diag.save()
    else:
        InternalDiagnosis.objects.create(plan=new_plan)
        
    # 3. Copiar Filosofía
    phil = getattr(original_plan, 'corporate_philosophy', None)
    if phil:
        phil.pk = None
        phil.plan = new_plan
        phil.save()
    else:
        CorporatePhilosophy.objects.create(plan=new_plan)
        
    # 4. Copiar Business Model Canvas
    for canvas in original_plan.canvas.all():
        canvas.pk = None
        canvas.plan = new_plan
        canvas.save()
        
    # 5. Copiar Matrices (FODA, EFI, EFE, MPC, etc)
    for matrix in original_plan.matrices.all():
        matrix.pk = None
        matrix.plan = new_plan
        matrix.save()
        
    messages.success(request, f'Plan "{original_plan.name}" clonado exitosamente.')
    return redirect('strategic_risk:dashboard')

@login_required
@check_module_access('Estrategia y Objetivos')
def plan_delete(request, pk):
    plan = get_object_or_404(StrategicPlan, pk=pk)
    
    if request.method == 'POST':
        plan_name = plan.name
        plan.delete()
        
        # If the deleted plan was the active one, clear the session
        if request.session.get('active_strategic_plan_id') == pk:
            del request.session['active_strategic_plan_id']
            
        messages.success(request, f'Plan "{plan_name}" eliminado correctamente.')
        return redirect('strategic_risk:dashboard')
        
    # Fallback to redirect if accessed via GET accidentally
    return redirect('strategic_risk:dashboard')

@login_required
def set_active_plan(request, pk):
    plan = get_object_or_404(StrategicPlan, pk=pk)
    request.session['active_strategic_plan_id'] = plan.id
    return redirect('strategic_risk:strat_data')

@login_required
@check_module_access('Estrategia y Objetivos')
def strat_data(request):
    plan_id = request.session.get('active_strategic_plan_id')
    if plan_id:
        plan = StrategicPlan.objects.filter(id=plan_id).first()
    else:
        plan = StrategicPlan.objects.order_by('-start_year').first()
        
    active_tab = 'externo'
    
    if plan:
        ext_env = getattr(plan, 'external_environment', None)
        fin_env = getattr(plan, 'financial_environment', None)
        int_diag = getattr(plan, 'internal_diagnosis', None)

        if request.method == 'POST':
            if 'save_external' in request.POST:
                ext_form = ExternalEnvironmentForm(request.POST, instance=ext_env)
                if ext_form.is_valid():
                    ext_form.save()
                    messages.success(request, 'Entorno Externo actualizado correctamente.')
                active_tab = 'externo'
            elif 'save_financial' in request.POST:
                fin_form = FinancialEnvironmentForm(request.POST, instance=fin_env)
                if fin_form.is_valid():
                    fin_form.save()
                    messages.success(request, 'Entorno Financiero actualizado correctamente.')
                active_tab = 'financiero'
            elif 'save_internal' in request.POST:
                int_form = InternalDiagnosisForm(request.POST, instance=int_diag)
                if int_form.is_valid():
                    int_form.save()
                    messages.success(request, 'Diagnóstico Interno actualizado correctamente.')
                active_tab = 'interno'

        ext_form = ExternalEnvironmentForm(instance=ext_env)
        fin_form = FinancialEnvironmentForm(instance=fin_env)
        int_form = InternalDiagnosisForm(instance=int_diag)
    else:
        ext_form = fin_form = int_form = None

    context = {
        'page_title': 'Planificación Estratégica - Diagnóstico y Entorno',
        'plan': plan,
        'ext_form': ext_form,
        'fin_form': fin_form,
        'int_form': int_form,
        'active_tab': active_tab
    }
    return render(request, 'strategic_risk/strat_data.html', context)

@login_required
@check_module_access('Estrategia y Objetivos')
def methodologies(request):
    plan_id = request.session.get('active_strategic_plan_id')
    if plan_id:
        plan = StrategicPlan.objects.filter(id=plan_id).first()
    else:
        plan = StrategicPlan.objects.order_by('-start_year').first()
        
    if not plan:
        messages.warning(request, "Debe seleccionar o crear un plan estratégico primero.")
        return redirect('strategic_risk:dashboard')

    foda = StrategicMatrix.objects.filter(plan=plan, matrix_type='FODA').first()
    foda = StrategicMatrix.objects.filter(plan=plan, matrix_type='FODA').first()
    foda_data = foda.data if foda and foda.data else {
        "F": [], "D": [], "O": [], "A": [], "FO": "", "DO": "", "FA": "", "DA": ""
    }

    efi = StrategicMatrix.objects.filter(plan=plan, matrix_type='EFI').first()
    efi_data = efi.data if efi and efi.data and isinstance(efi.data, list) else []

    efe = StrategicMatrix.objects.filter(plan=plan, matrix_type='EFE').first()
    efe_data = efe.data if efe and efe.data and isinstance(efe.data, list) else []

    mpc = StrategicMatrix.objects.filter(plan=plan, matrix_type='MPC').first()
    mpc_data = mpc.data if mpc and mpc.data and isinstance(mpc.data, dict) and mpc.data.get('factors') else {
        "competitors": ["Nuestra Entidad", "Competidor 1", "Competidor 2"],
        "factors": []
    }

    canvas_actual = BusinessModelCanvas.objects.filter(plan=plan, version__in=['1.0', 'Actual']).first()
    if canvas_actual and canvas_actual.version == '1.0':
        canvas_actual.version = 'Actual'
        canvas_actual.save()
    if not canvas_actual:
        canvas_actual = BusinessModelCanvas(plan=plan, version='Actual')
        
    canvas_futuro = BusinessModelCanvas.objects.filter(plan=plan, version='Futuro').first()
    if not canvas_futuro:
        canvas_futuro = BusinessModelCanvas(plan=plan, version='Futuro')
        canvas_futuro.save()
        
    philosophy, _ = CorporatePhilosophy.objects.get_or_create(plan=plan)
    
    conclusiones = StrategicMatrix.objects.filter(plan=plan, matrix_type='CONCLUSIONES').first()
    conclusiones_data = conclusiones.data if conclusiones and conclusiones.data else {
        "interno": [""] * 10,
        "externo": [""] * 10
    }
    
    context = {
        'page_title': 'Planificación Estratégica - Matrices Estratégicas',
        'plan': plan,
        'foda_data': json.dumps(foda_data),
        'efi_data': json.dumps(efi_data),
        'efe_data': json.dumps(efe_data),
        'mpc_data': json.dumps(mpc_data),
        'conclusiones_data': json.dumps(conclusiones_data),
        'canvas_actual': canvas_actual,
        'canvas_futuro': canvas_futuro,
        'philosophy': philosophy
    }
    return render(request, 'strategic_risk/methodologies.html', context)

@login_required
def save_matrix(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            plan_id = data.get('plan_id')
            matrix_type = data.get('matrix_type')
            matrix_data = data.get('data')
            
            plan = get_object_or_404(StrategicPlan, id=plan_id)
            matrix, created = StrategicMatrix.objects.get_or_create(plan=plan, matrix_type=matrix_type)
            matrix.data = matrix_data
            matrix.save()
            
            return JsonResponse({'status': 'success', 'message': f'Matriz {matrix_type} guardada correctamente.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
def save_canvas(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            plan_id = data.get('plan_id')
            version = data.get('version', 'Actual')
            
            plan = get_object_or_404(StrategicPlan, id=plan_id)
            canvas, created = BusinessModelCanvas.objects.get_or_create(plan=plan, version=version)
            
            canvas.key_partners = data.get('key_partners', '')
            canvas.key_activities = data.get('key_activities', '')
            canvas.key_resources = data.get('key_resources', '')
            canvas.value_proposition = data.get('value_proposition', '')
            canvas.customer_relationships = data.get('customer_relationships', '')
            canvas.channels = data.get('channels', '')
            canvas.customer_segments = data.get('customer_segments', '')
            canvas.cost_structure = data.get('cost_structure', '')
            canvas.revenue_streams = data.get('revenue_streams', '')
            canvas.save()
            
            return JsonResponse({'status': 'success', 'message': f'Business Model Canvas ({version}) guardado correctamente.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
def save_mpc(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            plan_id = data.get('plan_id')
            matrix_data = data.get('data')
            
            plan = get_object_or_404(StrategicPlan, id=plan_id)
            matrix, created = StrategicMatrix.objects.get_or_create(plan=plan, matrix_type='MPC')
            matrix.data = matrix_data
            matrix.save()
            
            return JsonResponse({'status': 'success', 'message': 'Matriz Perfil Competitivo (MPC) guardada correctamente.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
def save_metas_planeadas(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            plan_id = data.get('plan_id')
            metas_data = data.get('data')
            
            plan = get_object_or_404(StrategicPlan, id=plan_id)
            matrix, created = StrategicMatrix.objects.get_or_create(plan=plan, matrix_type='METAS')
            matrix.data = metas_data
            matrix.save()
            
            organization = getattr(request.user, 'organization', None)
            if not organization:
                from users.models import Organization
                organization = Organization.objects.first()
                if not organization:
                    organization = Organization.objects.create(name="Organización Principal")
            
            for row in metas_data:
                persp_name = row.get('perspectiva', '').upper().replace('/', '_').replace(' ', '_')
                
                tipo_persp = 'FINANCIERA'
                if 'CLIENTE' in persp_name or 'SOCIO' in persp_name:
                    tipo_persp = 'SOCIOS_CLIENTES'
                elif 'PROCESO' in persp_name:
                    tipo_persp = 'PROCESOS'
                elif 'APRENDIZAJE' in persp_name:
                    tipo_persp = 'APRENDIZAJE'

                perspectiva, _ = Perspectiva.objects.get_or_create(
                    organization=organization,
                    nombre=tipo_persp,
                    defaults={'plan': plan}
                )
                
                obj_nombre = row.get('objetivo', 'Objetivo sin nombre')
                tipo_obj = row.get('tipo', 'Estratégico')
                area_resp = row.get('area', '')
                responsable = row.get('responsable', '')
                
                objetivo, created_obj = ObjetivoEstrategico.objects.get_or_create(
                    organization=organization,
                    perspectiva=perspectiva,
                    nombre=obj_nombre,
                    defaults={
                        'codigo': f'OBJ-{ObjetivoEstrategico.objects.filter(organization=organization).count() + 1}',
                        'tipo_objetivo': tipo_obj,
                        'area_responsable': area_resp,
                        'responsable': responsable
                    }
                )
                
                if not created_obj:
                    objetivo.tipo_objetivo = tipo_obj
                    objetivo.area_responsable = area_resp
                    objetivo.responsable = responsable
                    objetivo.save()
                
                ind_nombre = row.get('indicador', 'Indicador')
                linea_base = row.get('base', '0').replace('%', '').strip()
                try: linea_base = float(linea_base)
                except: linea_base = 0

                indicador, _ = Indicador.objects.get_or_create(
                    organization=organization,
                    objetivo=objetivo,
                    nombre=ind_nombre,
                    defaults={
                        'unidad_medida': '%',
                        'frecuencia_medicion': 'ANUAL',
                        'linea_base': linea_base
                    }
                )
                
                for i in range(1, 4):
                    meta_val = row.get(f'meta{i}', '0').replace('%', '').strip()
                    try: meta_val = float(meta_val)
                    except: continue
                    
                    meta, _ = MetaPeriodo.objects.update_or_create(
                        indicador=indicador,
                        periodo=f'Meta {i}',
                        defaults={'meta_programada': meta_val}
                    )
            
            return JsonResponse({'status': 'success', 'message': 'Metas Planeadas guardadas correctamente.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
def save_philosophy(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            plan_id = request.session.get('active_strategic_plan_id')
            if not plan_id:
                plan_id = StrategicPlan.objects.order_by('-start_year').first().id
                
            plan = get_object_or_404(StrategicPlan, id=plan_id)
            philosophy, _ = CorporatePhilosophy.objects.get_or_create(plan=plan)
            
            field_type = data.get('type')
            text = data.get('text', '')
            
            if field_type == 'mision':
                philosophy.mission = text
                msg = 'Misión guardada correctamente.'
            elif field_type == 'vision':
                philosophy.vision = text
                msg = 'Visión guardada correctamente.'
            elif field_type == 'valores':
                philosophy.values = text
                msg = 'Valores guardados correctamente.'
            else:
                return JsonResponse({'status': 'error', 'message': 'Tipo inválido.'})
                
            philosophy.save()
            return JsonResponse({'status': 'success', 'message': msg})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
@check_module_access('Estrategia y Objetivos')
def controls(request):
    plan_id = request.session.get('active_strategic_plan_id')
    if plan_id:
        plan = StrategicPlan.objects.filter(id=plan_id).first()
    else:
        plan = StrategicPlan.objects.order_by('-start_year').first()
        
    metas_matrix = StrategicMatrix.objects.filter(plan=plan, matrix_type='METAS').first() if plan else None
    
    objetivos_ingresados = ObjetivoEstrategico.objects.filter(perspectiva__plan=plan) if plan else []
    indicadores_ingresados = Indicador.objects.filter(objetivo__perspectiva__plan=plan) if plan else []
    
    context = {
        'page_title': 'Planificación Estratégica - FASE ESTRATEGICA',
        'plan': plan,
        'perspectivas': plan.perspectivas.all() if plan else [],
        'tipos_objetivo': plan.tipos_objetivo.all() if plan else [],
        'areas_responsables': plan.areas_responsables.all() if plan else [],
        'responsables': plan.responsables.all() if plan else [],
        'objetivos_ingresados': objetivos_ingresados,
        'indicadores_ingresados': indicadores_ingresados,
    }
    return render(request, 'strategic_risk/controls.html', context)

@login_required
def add_objective(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            objective_id = data.get('objective_id')
            perspective_id = data.get('perspective_id')
            nombre = data.get('name')
            descripcion = data.get('description', '')
            propuesta_valor = data.get('propuesta_valor', '')
            tipo_objetivo = data.get('tipo_objetivo', '')
            area_responsable = data.get('area_responsable', '')
            responsable = data.get('responsable', '')
            
            if propuesta_valor:
                descripcion = f"{descripcion}\n\nPropuesta de Valor:\n{propuesta_valor}".strip()
            
            perspectiva = get_object_or_404(Perspectiva, id=perspective_id)
            
            organization = getattr(request.user, 'organization', None)
            if not organization:
                from users.models import Organization
                organization = Organization.objects.first()
                
            if objective_id:
                obj = get_object_or_404(ObjetivoEstrategico, id=objective_id, organization=organization)
                obj.perspectiva = perspectiva
                obj.nombre = nombre
                obj.descripcion = descripcion
                obj.tipo_objetivo = tipo_objetivo
                obj.area_responsable = area_responsable
                obj.responsable = responsable
                obj.save()
                message = 'Objetivo actualizado correctamente.'
                obj_id = obj.id
            else:
                codigo = f"OE-{ObjetivoEstrategico.objects.filter(organization=organization).count() + 1}"
                obj = ObjetivoEstrategico.objects.create(
                    organization=organization,
                    perspectiva=perspectiva, 
                    nombre=nombre, 
                    descripcion=descripcion,
                    tipo_objetivo=tipo_objetivo,
                    area_responsable=area_responsable,
                    responsable=responsable,
                    codigo=codigo
                )
                message = 'Objetivo añadido correctamente.'
                obj_id = obj.id
            
            return JsonResponse({'status': 'success', 'message': message, 'id': obj_id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
def delete_objective(request, pk):
    if request.method == 'POST':
        organization = getattr(request.user, 'organization', None)
        if not organization:
            from users.models import Organization
            organization = Organization.objects.first()
        obj = get_object_or_404(ObjetivoEstrategico, id=pk, organization=organization)
        obj.delete()
        return JsonResponse({'status': 'success', 'message': 'Objetivo eliminado correctamente.'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
def add_kpi(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            objective_id = data.get('objective_id')
            nombre = data.get('nombre')
            formula = data.get('formula', '')
            peso = data.get('peso', 0)
            unidad_medida = data.get('unidad_medida', 'Porcentaje')
            frecuencia_medicion = data.get('frecuencia_medicion', 'MENSUAL')
            fecha_inicio = data.get('fecha_inicio') or None
            fecha_fin = data.get('fecha_fin') or None
            responsable = data.get('responsable', '')
            medio_verificacion = data.get('medio_verificacion', '')
            tipo_objetivo = data.get('tipo_objetivo', '')
            
            organization = getattr(request.user, 'organization', None)
            if not organization:
                from users.models import Organization
                organization = Organization.objects.first()

            objective = get_object_or_404(ObjetivoEstrategico, id=objective_id)
            
            kpi_id = data.get('id')
            if kpi_id:
                kpi = Indicador.objects.get(id=kpi_id, organization=organization)
                kpi.nombre = nombre
                kpi.formula = formula
                kpi.peso = peso
                kpi.unidad_medida = unidad_medida
                kpi.frecuencia_medicion = frecuencia_medicion
                kpi.fecha_inicio = fecha_inicio
                kpi.fecha_fin = fecha_fin
                kpi.responsable = responsable
                kpi.medio_verificacion = medio_verificacion
                kpi.tipo_objetivo = tipo_objetivo
                kpi.save()
                message = 'Indicador actualizado correctamente.'
            else:
                kpi = Indicador.objects.create(
                    organization=organization,
                    objetivo=objective, 
                    nombre=nombre, 
                    formula=formula,
                    peso=peso,
                    unidad_medida=unidad_medida,
                    frecuencia_medicion=frecuencia_medicion,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    responsable=responsable,
                    medio_verificacion=medio_verificacion,
                    tipo_objetivo=tipo_objetivo
                )
                message = 'Indicador añadido correctamente.'
            
            return JsonResponse({'status': 'success', 'message': message})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
@check_module_access('Estrategia y Objetivos')
def delete_kpi(request, pk):
    if request.method == 'POST':
        try:
            organization = getattr(request.user, 'organization', None)
            if not organization:
                from users.models import Organization
                organization = Organization.objects.first()

            kpi = Indicador.objects.get(id=pk, organization=organization)
            kpi.delete()
            return JsonResponse({'status': 'success', 'message': 'Indicador eliminado exitosamente'})
        except Indicador.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'El indicador no existe o no tiene permisos.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'})

@login_required
@check_module_access('Estrategia y Objetivos')
def delete_multiple_kpis(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            kpi_ids = data.get('kpi_ids', [])
            if not kpi_ids:
                return JsonResponse({'status': 'error', 'message': 'No se seleccionaron indicadores'})

            organization = getattr(request.user, 'organization', None)
            if not organization:
                from users.models import Organization
                organization = Organization.objects.first()

            deleted_count, _ = Indicador.objects.filter(id__in=kpi_ids, organization=organization).delete()
            return JsonResponse({'status': 'success', 'message': f'{deleted_count} indicadores eliminados exitosamente'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'})

@login_required
@check_module_access('Estrategia y Objetivos')
def save_ponderaciones(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ponderaciones = data.get('ponderaciones', [])
            
            for item in ponderaciones:
                obj_id = item.get('id')
                peso = item.get('peso')
                try:
                    obj = ObjetivoEstrategico.objects.get(id=obj_id)
                    obj.peso = peso
                    obj.save()
                except ObjetivoEstrategico.DoesNotExist:
                    continue
            
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
@check_module_access('Estrategia y Objetivos')
def reports(request):
    plan_id = request.session.get('active_strategic_plan_id')
    if plan_id:
        plan = StrategicPlan.objects.filter(id=plan_id).first()
    else:
        plan = StrategicPlan.objects.order_by('-start_year').first()
    
    context = {
        'page_title': 'Planificación Estratégica - Reporte Final',
        'plan': plan
    }
    
    if plan:
        context['ext_env'] = getattr(plan, 'external_environment', None)
        context['fin_env'] = getattr(plan, 'financial_environment', None)
        context['int_diag'] = getattr(plan, 'internal_diagnosis', None)
        context['canvas'] = plan.canvas.filter(version='Futuro').first() or plan.canvas.filter(version='Actual').first()
        
        foda = plan.matrices.filter(matrix_type='FODA').first()
        if foda and foda.data:
            normalized_foda = {}
            for q in ['fortalezas', 'debilidades', 'oportunidades', 'amenazas']:
                val = foda.data.get(q, [])
                if isinstance(val, str):
                    normalized_foda[q] = [{'text': v.lstrip('0123456789. ').strip(), 'weight': 5} for v in val.split('\n') if v.strip()]
                else:
                    normalized_foda[q] = val
            
            # Calculate metrics for the combined view
            # Internal
            total_internal_weight = sum(int(item.get('weight', 0)) for item in normalized_foda.get('fortalezas', []) + normalized_foda.get('debilidades', []))
            for item in normalized_foda.get('fortalezas', []):
                w = int(item.get('weight', 0))
                item['pct'] = round((w / total_internal_weight * 100) if total_internal_weight else 0, 1)
                item['rating_label'] = 'MF' if w >= 8 else ('F' if w >= 5 else 'M')
                item['rating_val'] = 4 if w >= 8 else (3 if w >= 5 else 2)
                item['score'] = round((item['pct'] / 100) * item['rating_val'], 2)
                
            mefi_total = 0
            for item in normalized_foda.get('debilidades', []):
                w = int(item.get('weight', 0))
                item['pct'] = round((w / total_internal_weight * 100) if total_internal_weight else 0, 1)
                item['rating_label'] = 'MD' if w >= 8 else ('D' if w >= 5 else 'M')
                item['rating_val'] = 1 if w >= 8 else (2 if w >= 5 else 3)
                item['score'] = round((item['pct'] / 100) * item['rating_val'], 2)
                mefi_total += item['score']

            for item in normalized_foda.get('fortalezas', []):
                mefi_total += item['score']

            normalized_foda['mefi_total'] = round(mefi_total, 2)
            if normalized_foda['mefi_total'] >= 2.5:
                normalized_foda['mefi_analisis'] = f"El puntaje total ponderado de la matriz MEFI es de {normalized_foda['mefi_total']}. Este resultado, al ser mayor que el promedio de 2.5, indica que la organización posee una posición interna fuerte. En general, las fortalezas superan a las debilidades, lo que sugiere que la entidad está capitalizando sus recursos y ventajas internas de forma adecuada."
            else:
                normalized_foda['mefi_analisis'] = f"El puntaje total ponderado de la matriz MEFI es de {normalized_foda['mefi_total']}. Este resultado, al ser menor al promedio de 2.5, revela una posición interna débil. Las debilidades organizacionales están afectando el desempeño general, requiriendo acciones correctivas inmediatas para mejorar las capacidades operativas y estratégicas."

            # External
            total_external_weight = sum(int(item.get('weight', 0)) for item in normalized_foda.get('oportunidades', []) + normalized_foda.get('amenazas', []))
            for item in normalized_foda.get('oportunidades', []):
                w = int(item.get('weight', 0))
                item['pct'] = round((w / total_external_weight * 100) if total_external_weight else 0, 1)
                item['rating_label'] = 'MF' if w >= 8 else ('F' if w >= 5 else 'M')
                item['rating_val'] = 4 if w >= 8 else (3 if w >= 5 else 2)
                item['score'] = round((item['pct'] / 100) * item['rating_val'], 2)

            mefe_total = 0
            for item in normalized_foda.get('amenazas', []):
                w = int(item.get('weight', 0))
                item['pct'] = round((w / total_external_weight * 100) if total_external_weight else 0, 1)
                item['rating_label'] = 'MD' if w >= 8 else ('D' if w >= 5 else 'M')
                item['rating_val'] = 1 if w >= 8 else (2 if w >= 5 else 3)
                item['score'] = round((item['pct'] / 100) * item['rating_val'], 2)
                mefe_total += item['score']
                
            for item in normalized_foda.get('oportunidades', []):
                mefe_total += item['score']

            normalized_foda['mefe_total'] = round(mefe_total, 2)
            if normalized_foda['mefe_total'] >= 2.5:
                normalized_foda['mefe_analisis'] = f"El puntaje total ponderado de la matriz MEFE es de {normalized_foda['mefe_total']}. Al estar por encima de la media (2.5), se deduce que la entidad está respondiendo favorablemente a su entorno. Las estrategias actuales aprovechan eficazmente las oportunidades existentes mientras mitigan los efectos adversos de las amenazas externas."
            else:
                normalized_foda['mefe_analisis'] = f"El puntaje total ponderado de la matriz MEFE es de {normalized_foda['mefe_total']}. Este valor, por debajo de la media (2.5), indica que la entidad no está respondiendo eficientemente a los factores externos. Las estrategias actuales deben replantearse para capitalizar mejor las oportunidades y crear un escudo robusto ante las amenazas del entorno."

            normalized_foda['estrategias_fo'] = foda.data.get('estrategias_fo', '')
            normalized_foda['estrategias_fa'] = foda.data.get('estrategias_fa', '')
            normalized_foda['estrategias_do'] = foda.data.get('estrategias_do', '')
            normalized_foda['estrategias_da'] = foda.data.get('estrategias_da', '')

            foda.data = normalized_foda
        context['foda'] = foda
        
        mpc = plan.matrices.filter(matrix_type='MPC').first()
        if mpc and mpc.data and 'factors' in mpc.data:
            comps = mpc.data.get('competitors', [])
            factors = mpc.data.get('factors', [])
            num_comps = len(comps)
            totals = [0] * (num_comps + 1)
            
            for factor in factors:
                w = factor.get('weight', 0)
                factor['weight_pct'] = round(w * 100, 1)
                ratings = factor.get('ratings', [])
                scores = []
                for i in range(num_comps + 1):
                    r = ratings[i] if i < len(ratings) else 0
                    score = round(w * r, 2)
                    scores.append({'rating': r, 'score': score})
                    totals[i] += score
                factor['scores'] = scores
                
            mpc.data['totals'] = [round(t, 2) for t in totals]
            
        context['mpc'] = mpc

    return render(request, 'strategic_risk/reports.html', context)

import openpyxl
from django.http import HttpResponse

@login_required
def export_bsc_excel(request):
    organization = request.user.organization if hasattr(request.user, 'organization') else None
    if not organization:
        return HttpResponse("Organización no encontrada.", status=400)
        
    wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = "FASE ESTRATEGICA"
    
    # Header
    headers = ["Perspectiva", "Objetivo", "Indicador", "Línea Base", "Periodo Meta", "Meta Programada", "Resultado Real", "% Cumplimiento", "Semáforo"]
    ws.append(headers)
    
    # Formato Header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
    # Obtener Metas
    from .models import MetaPeriodo
    metas = MetaPeriodo.objects.filter(indicador__organization=organization).select_related(
        'indicador', 'indicador__objetivo', 'indicador__objetivo__perspectiva'
    ).order_by('indicador__objetivo__perspectiva__nombre', 'indicador__objetivo__codigo', 'periodo')
    
    for meta in metas:
        indicador = meta.indicador
        objetivo = indicador.objetivo
        perspectiva = objetivo.perspectiva
        
        ws.append([
            perspectiva.nombre,
            f"{objetivo.codigo} - {objetivo.nombre}",
            indicador.nombre,
            float(indicador.linea_base) if indicador.linea_base else 0,
            meta.periodo,
            float(meta.meta_programada) if meta.meta_programada else 0,
            float(meta.resultado_real) if meta.resultado_real else None,
            float(meta.porcentaje_cumplimiento) if meta.porcentaje_cumplimiento else None,
            meta.semaforo
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Balanced_Scorecard.xlsx"'
    wb.save(response)
    return response

# --- VISTA PÚBLICA PARA ENCUESTAS ---
def public_survey(request, survey_id):
    """
    Vista pública para que cualquier persona con el enlace pueda responder una encuesta.
    No requiere autenticación.
    """
    context = {
        'survey_id': survey_id,
        # Si hubiera un modelo de encuesta, se buscaría aquí y se pasaría al contexto
    }
    return render(request, 'strategic_risk/public_survey.html', context)


@login_required
@require_http_methods(["GET"])
def get_kpi_metas(request, pk):
    try:
        from .models import MetaPeriodo
        ind = get_object_or_404(Indicador, pk=pk)
        
        # Verify permissions
        org = getattr(request.user, 'organization', None)
        if ind.organization != org:
            return JsonResponse({'status': 'error', 'message': 'No tiene permisos para acceder a este indicador.'}, status=403)
            
        metas = MetaPeriodo.objects.filter(indicador=ind).order_by('periodo')
        metas_list = [{
            'id': m.id,
            'periodo': m.periodo,
            'meta_programada': float(m.meta_programada) if m.meta_programada is not None else 0.0,
            'resultado_real': float(m.resultado_real) if m.resultado_real is not None else None,
            'porcentaje_cumplimiento': float(m.porcentaje_cumplimiento) if m.porcentaje_cumplimiento is not None else None,
        } for m in metas]
        
        return JsonResponse({
            'status': 'success',
            'indicador': {
                'id': ind.id,
                'perspectiva': ind.objetivo.perspectiva.nombre if ind.objetivo and ind.objetivo.perspectiva else '',
                'tipo_objetivo': ind.tipo_objetivo or (ind.objetivo.tipo_objetivo if ind.objetivo else ''),
                'objetivo': ind.objetivo.nombre if ind.objetivo else '',
                'nombre': ind.nombre,
                'formula': ind.formula,
                'unidad_medida': ind.unidad_medida,
                'linea_base': str(ind.linea_base) if ind.linea_base else '0.00',
                'fecha_inicio': ind.fecha_inicio.strftime('%Y-%m-%d') if ind.fecha_inicio else '',
                'fecha_fin': ind.fecha_fin.strftime('%Y-%m-%d') if ind.fecha_fin else '',
                'frecuencia': ind.get_frecuencia_medicion_display(),
                'responsable': ind.responsable,
            },
            'metas': metas_list
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def save_kpi_metas(request, pk):
    try:
        import json
        from .models import MetaPeriodo
        ind = get_object_or_404(Indicador, pk=pk)
        
        # Verify permissions
        org = getattr(request.user, 'organization', None)
        if ind.organization != org:
            return JsonResponse({'status': 'error', 'message': 'No tiene permisos para modificar este indicador.'}, status=403)
            
        data = json.loads(request.body)
        metas_data = data.get('metas', [])
        
        # Delete existing metas for this indicator to replace with new ones (or update them)
        # We can update if ID is provided, else create. Or just clear and recreate?
        # Actually, simpler to just clear and recreate if they are sending all metas.
        # But wait, 'resultado_real' etc might be lost if we just delete!
        # Let's update or create by periodo
        
        existing_metas = MetaPeriodo.objects.filter(indicador=ind)
        existing_periodos = [m.periodo for m in existing_metas]
        
        new_periodos = []
        for m_data in metas_data:
            periodo = m_data.get('periodo')
            if not periodo: continue
            new_periodos.append(periodo)
            meta_val = m_data.get('meta_programada', 0)
            
            # Update or create
            meta_obj, created = MetaPeriodo.objects.update_or_create(
                indicador=ind,
                periodo=periodo,
                defaults={
                    'meta_programada': meta_val,
                    'organization': org
                }
            )
            
        # Optional: remove metas that are no longer in the payload?
        # Let's just keep them or delete ones not in payload
        # MetaPeriodo.objects.filter(indicador=ind).exclude(periodo__in=new_periodos).delete()
        
        return JsonResponse({'status': 'success', 'message': 'Metas guardadas correctamente.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
